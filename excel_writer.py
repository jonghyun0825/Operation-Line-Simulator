"""Lot별 실제 .xlsx 파일 기록. 저장 실패(파일 잠금 등) 시에도 서버가 죽지 않도록 한다.

워크북 객체는 메모리에 유지되므로, wb.save()가 한 번 실패해도 데이터는 유실되지 않는다.
다음 기록 시점에 다시 save()가 호출되면 그 사이 누적된 내용까지 함께 저장된다
(별도의 재시도 큐 자료구조 없이 in-memory workbook 자체가 큐 역할을 한다).
저장 위치는 config.EXCEL_DIR(OneDrive 동기화 경로일 수 있음)이므로 이 재시도 동작이 특히 중요하다.
"""
import os
from typing import Dict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import config

HEADER_FONT = Font(name="Arial", bold=True)
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
BODY_FONT = Font(name="Arial")
NG_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
TITLE_FONT = Font(name="Arial", bold=True, size=14)
SPEC_FONT = Font(name="Arial", italic=True, color="595959")

# 측정기록 시트: 반제품(SN)당 1행. 1행=규격 안내, 2행=컬럼 헤더, 3행부터 데이터.
MEASURE_HEADERS = ["순번", "반제품SN", "토크1(ST1)", "판정1", "ST1시각", "토크2(ST2)", "판정2", "ST2시각", "최종판정"]
MEASURE_HEADER_ROW = 2
MEASURE_FIRST_DATA_ROW = 3
COL_SEQ, COL_SN = 1, 2
COL_VALUE = {"ST1": 3, "ST2": 6}
COL_JUDGMENT = {"ST1": 4, "ST2": 7}
COL_TS = {"ST1": 5, "ST2": 8}
COL_FINAL = 9

NG_HEADERS = ["반제품SN", "NG공정", "측정값", "발생시각"]


class ExcelWriter:
    def __init__(self, lot_id: str):
        self.lot_id = lot_id
        os.makedirs(config.EXCEL_DIR, exist_ok=True)
        self.path = os.path.join(config.EXCEL_DIR, f"{lot_id}.xlsx")

        self.wb = Workbook()
        self.ws_measure = self.wb.active
        self.ws_measure.title = "측정기록"
        self._write_spec_row(self.ws_measure)
        self._write_header(self.ws_measure, MEASURE_HEADERS, row=MEASURE_HEADER_ROW)
        self._sn_row: Dict[str, int] = {}
        self._next_row = MEASURE_FIRST_DATA_ROW

        self.ws_ng = self.wb.create_sheet("NG추적")
        self._write_header(self.ws_ng, NG_HEADERS, row=1)

        self.ws_summary = self.wb.create_sheet("Lot요약")

        self._save()

    def _write_spec_row(self, ws):
        st1, st2 = config.ST1_SPEC, config.ST2_SPEC
        text = (
            f"규격 — ST1({st1['name']}): {st1['lsl']} ~ {st1['usl']} N·m   |   "
            f"ST2({st2['name']}): {st2['lsl']} ~ {st2['usl']} N·m"
        )
        ws.cell(row=1, column=1, value=text).font = SPEC_FONT

    def _write_header(self, ws, headers, row: int):
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

    def _save(self) -> bool:
        try:
            self.wb.save(self.path)
            return True
        except Exception as e:
            print(f"[excel_writer] '{self.path}' 저장 실패 (다음 기록 시 재시도): {e}")
            return False

    def append_measurement(self, sn: str, station: str, value: float, judgment: str, ts: str):
        """반제품(sn)당 1행 구조: ST1 측정이 새 행을 만들고, ST2 측정이 같은 행을 채운 뒤
        두 판정이 모두 존재하면 최종판정을 계산한다 (호출 순서에 관계없이 동작)."""
        ws = self.ws_measure
        if sn not in self._sn_row:
            row = self._next_row
            self._next_row += 1
            self._sn_row[sn] = row
            seq = len(self._sn_row)
            ws.cell(row=row, column=COL_SEQ, value=seq).font = BODY_FONT
            ws.cell(row=row, column=COL_SN, value=sn).font = BODY_FONT
        row = self._sn_row[sn]

        value_cell = ws.cell(row=row, column=COL_VALUE[station], value=value)
        value_cell.font = BODY_FONT
        if judgment == "NG":
            value_cell.fill = NG_FILL
        ws.cell(row=row, column=COL_JUDGMENT[station], value=judgment).font = BODY_FONT
        ws.cell(row=row, column=COL_TS[station], value=ts).font = BODY_FONT

        j1 = ws.cell(row=row, column=COL_JUDGMENT["ST1"]).value
        j2 = ws.cell(row=row, column=COL_JUDGMENT["ST2"]).value
        if j1 and j2:
            final = "NG" if "NG" in (j1, j2) else "OK"
            ws.cell(row=row, column=COL_FINAL, value=final).font = BODY_FONT

        if judgment == "NG":
            self._append_ng_row(sn, station, value, ts)

        self._save()

    def _append_ng_row(self, sn, station, value, ts):
        row = self.ws_ng.max_row + 1
        for col, v in enumerate([sn, station, value, ts], start=1):
            cell = self.ws_ng.cell(row=row, column=col, value=v)
            cell.font = BODY_FONT

    def write_summary(self, agg: dict):
        ws = self.ws_summary
        ws.cell(row=1, column=1, value="Lot 요약").font = TITLE_FONT

        info_rows = [
            ("Lot ID", agg["lot_id"]),
            ("생성시각", agg["created_at"]),
            ("완료시각", agg.get("completed_at", "")),
            ("온도(°C)", agg["temperature"]),
            ("헤드 속도", agg["head_speed"]),
            ("박스 수", agg["boxes"]),
            ("생산수량", agg["quantity"]),
            ("OK수", agg["ok_count"]),
            ("NG수", agg["ng_count"]),
            ("불량률(%)", agg["defect_rate"]),
        ]
        r = 3
        for label, value in info_rows:
            ws.cell(row=r, column=1, value=label).font = HEADER_FONT
            ws.cell(row=r, column=1).fill = HEADER_FILL
            ws.cell(row=r, column=2, value=value).font = BODY_FONT
            r += 1

        r += 1
        headers = ["공정", "측정항목", "평균", "표준편차", "최소", "최대", "NG수"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        r += 1
        for station, stat in agg["stations"].items():
            values = [station, stat["name"], stat["mean"], stat["std"], stat["min"], stat["max"], stat["ng_count"]]
            for col, v in enumerate(values, start=1):
                ws.cell(row=r, column=col, value=v).font = BODY_FONT
            r += 1

        for ws_ in (self.ws_measure, self.ws_ng, self.ws_summary):
            for col in range(1, 10):
                ws_.column_dimensions[get_column_letter(col)].width = 16

        self._save()
